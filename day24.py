#Arkusze kalkulacyjne
import openpyxl, os, pprint

path='C:\\Users\\ASUS\\Desktop\\DANE 03\\segregator\\programowanie\\pythonProject\\data science Python'
os.getcwd() #ustawienie katalogu roboczego
os.chdir(path) #zmiana katalogu roboczego
openXLSX=openpyxl.load_workbook('data.xlsx') #wczytanie skoroszytu
print(type(openXLSX)) #zwraca typ: <class 'openpyxl.workbook.workbook.Workbook'>
print(openXLSX.get_sheet_names()) #zwraca wszystkie arkusze w skoroszycie
dataSheet=openXLSX.get_sheet_by_name('Arkusz3')
print(dataSheet) #zwraca dany arkusz
print(dataSheet.title) #zwraca tytuł danego arkusza
dataSheet2=openXLSX.get_sheet_by_name('Arkusz1')
print(dataSheet2['A1'].value) #zwraca zawartość danej komórki: 2015-04-05 13:34:00
print('Komórka '+dataSheet2['B1'].coordinate+' ma wartość: '+dataSheet2['B1'].value)
print(dataSheet2.cell(row=1, column=2).value) #metoda cell również zwraca zawartość danej komórki w arkuszu: jabłka
for x in range(1,8,2):
    print(dataSheet2.cell(row=x, column=3).value) #wypisuje co drugą komórke z 3 kolumny podanego arkusza

krotka1=tuple(dataSheet2['A1':'C3'])
print(krotka1) #zwraca typy podanego zakresu komórek
#print(dataSheet2['A1':'C3'].value) #nie można wykonać bo: 'tuple' object has no attribute 'value'

for Rowcell in dataSheet2['A1':'C3']: 
    for cell in Rowcell:
        print(cell.coordinate, cell.value)
    print('--- KONIEC WIERSZA ---')

#bład zwiazany z typem generatora: "TypeError: 'generator' object is not subscriptable" - trzeba zamienić na liste i będzie GIT ;)
# print(dataSheet2.columns[1])
# for cell in dataSheet2.columns[1]:
#     print(cell.value)

#projekt: odczyt z arkusza danych i zapis w .py file
print('Otwieranie skoroszytu...')
wb=openpyxl.load_workbook(os.path.join(path, 'dataUnitedStates.xlsx'))
sheet=wb.get_sheet_by_name('USAdata')
countryData={}
print('Odczyt wierszy...')
for row in range(2, sheet.max_row+1):
    state=sheet['B'+str(row)].value
    country=sheet['C'+str(row)].value
    pop=sheet['D'+str(row)].value
    countryData.setdefault(state,{})
    countryData[state].setdefault(country, {'tract': 0, 'pop': 0})
    countryData[state][country]['tract']+=1
    countryData[state][country]['pop']+=int(pop)

print('Zapisywanie wyników...')
resultFile=open(os.path.join(path,'wynikiUSA.py'), 'w')
resultFile.write('allData='+pprint.pformat(countryData))
resultFile.close()
print('Gotowe!')